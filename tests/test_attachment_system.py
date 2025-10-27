"""
Test suite for attachment detection and processing system

Tests cover:
1. Popup window detection
2. Document parsing (PDF, XLSX, HWP)
3. Full extraction workflow
4. Metadata validation
"""

import pytest
import asyncio
from pathlib import Path
from unittest.mock import Mock, AsyncMock, patch
from datetime import datetime

from crawler.popup_handler import PopupWindowHandler
from crawler.document_parser import DocumentParser
from crawler.content_extractor import SeoulTrafficContentExtractor
from config.schemas import (
    AttachedDocument,
    AttachmentType,
    AttachmentSource,
    PageMetadata
)


class TestPopupWindowHandler:
    """Test PopupWindowHandler functionality"""

    @pytest.fixture
    def handler(self, tmp_path):
        """Create PopupWindowHandler instance"""
        return PopupWindowHandler(
            download_dir=tmp_path / "downloads",
            timeout=5000
        )

    @pytest.mark.asyncio
    async def test_detect_attachment_links_pdf(self, handler):
        """Test detection of PDF attachment links"""
        html = """
        <html>
            <body>
                <a href="document.pdf" target="_blank">PDF 문서</a>
                <a href="report.xlsx">엑셀 보고서</a>
                <a href="#" onclick="window.open('popup.hwp')">한글 문서</a>
            </body>
        </html>
        """

        # Mock page object
        mock_page = AsyncMock()
        mock_page.query_selector_all = AsyncMock(return_value=[])

        links = await handler.detect_attachment_links(mock_page, html)

        # Should detect 3 attachment links
        assert len(links) >= 2  # At least PDF and XLSX
        assert any(link['href'].endswith('.pdf') for link in links)
        assert any(link['href'].endswith('.xlsx') for link in links)

    @pytest.mark.asyncio
    async def test_popup_listener_setup(self, handler):
        """Test popup window listener setup"""
        mock_context = Mock()
        mock_page = Mock()

        await handler.setup_popup_listener(mock_context, mock_page)

        # Verify listener was registered
        assert mock_context.on.called


class TestDocumentParser:
    """Test DocumentParser functionality"""

    @pytest.fixture
    def parser(self, tmp_path):
        """Create DocumentParser instance"""
        from utils.entity_preview import SimpleEntityExtractor
        return DocumentParser(
            attachments_dir=tmp_path / "attachments",
            entity_extractor=SimpleEntityExtractor()
        )

    def test_create_sample_pdf(self, tmp_path):
        """Create a sample PDF for testing"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter

            pdf_path = tmp_path / "test_document.pdf"
            c = canvas.Canvas(str(pdf_path), pagesize=letter)
            c.drawString(100, 750, "Test PDF Document")
            c.drawString(100, 730, "지하철 2호선 운행 지연")
            c.drawString(100, 710, "강남역에서 신호 장애 발생")
            c.save()

            return pdf_path
        except ImportError:
            pytest.skip("reportlab not installed")

    @pytest.mark.asyncio
    async def test_parse_pdf(self, parser, tmp_path):
        """Test PDF parsing"""
        # Create sample PDF
        pdf_path = self.test_create_sample_pdf(tmp_path)

        # Create AttachedDocument
        attachment = AttachedDocument(
            attachment_id="test-001",
            attachment_type=AttachmentType.PDF,
            original_filename="test_document.pdf",
            file_size=pdf_path.stat().st_size,
            source_type=AttachmentSource.DIRECT_LINK,
            source_url="http://example.com/test.pdf",
            local_path=str(pdf_path),
            extraction_method="pypdf",
            extraction_success=False,
            extracted_text="",
            extracted_tables=[],
            extracted_images=[],
            word_count=0,
            entities_preview=[],
            named_entities={}
        )

        # Parse document
        parsed = await parser.parse_document(attachment)

        # Verify parsing
        assert parsed.extraction_success is True
        assert len(parsed.extracted_text) > 0
        assert parsed.word_count > 0
        assert "지하철" in parsed.extracted_text or "Test" in parsed.extracted_text

    @pytest.mark.asyncio
    async def test_parse_xlsx(self, parser, tmp_path):
        """Test XLSX parsing"""
        try:
            from openpyxl import Workbook

            # Create sample XLSX
            xlsx_path = tmp_path / "test_spreadsheet.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "교통 데이터"
            ws['A1'] = "노선"
            ws['B1'] = "구간"
            ws['C1'] = "지연시간"
            ws['A2'] = "2호선"
            ws['B2'] = "강남-선릉"
            ws['C2'] = "15분"
            wb.save(xlsx_path)

            # Create AttachedDocument
            attachment = AttachedDocument(
                attachment_id="test-002",
                attachment_type=AttachmentType.XLSX,
                original_filename="test_spreadsheet.xlsx",
                file_size=xlsx_path.stat().st_size,
                source_type=AttachmentSource.POPUP_WINDOW,
                source_url="http://example.com/test.xlsx",
                local_path=str(xlsx_path),
                extraction_method="openpyxl",
                extraction_success=False,
                extracted_text="",
                extracted_tables=[],
                extracted_images=[],
                word_count=0,
                entities_preview=[],
                named_entities={}
            )

            # Parse document
            parsed = await parser.parse_document(attachment)

            # Verify parsing
            assert parsed.extraction_success is True
            assert len(parsed.extracted_tables) > 0
            assert parsed.sheet_names == ["교통 데이터"]
            assert "2호선" in parsed.extracted_text

        except ImportError:
            pytest.skip("openpyxl not installed")


class TestContentExtractor:
    """Test full ContentExtractor workflow"""

    @pytest.fixture
    def extractor(self, tmp_path):
        """Create ContentExtractor instance"""
        return SeoulTrafficContentExtractor(
            output_dir=tmp_path / "output",
            download_dir=tmp_path / "downloads",
            headless=True,
            verbose=False
        )

    def test_directory_creation(self, extractor):
        """Test that necessary directories are created"""
        assert extractor.raw_html_dir.exists()
        assert extractor.markdown_dir.exists()
        assert extractor.metadata_dir.exists()
        assert extractor.attachments_dir.exists()

    @pytest.mark.asyncio
    async def test_build_metadata(self, extractor):
        """Test metadata building"""
        # Sample data
        url = "https://news.seoul.go.kr/traffic/article/123"
        html = """
        <html>
            <head><title>지하철 2호선 운행 지연 안내</title></head>
            <body>
                <article>
                    <p>강남역에서 신호 장애로 인한 지연이 발생했습니다.</p>
                    <a href="/traffic/category">교통 정보</a>
                </article>
            </body>
        </html>
        """
        markdown = "# 지하철 2호선 운행 지연 안내\n\n강남역에서 신호 장애로 인한 지연이 발생했습니다."
        links = ["https://news.seoul.go.kr/traffic/category"]
        url_tree = {
            url: {
                'parent_url': 'https://news.seoul.go.kr/traffic',
                'depth': 2
            }
        }
        attached_documents = []

        # Build metadata
        metadata = await extractor._build_metadata(
            url=url,
            html=html,
            markdown=markdown,
            links=links,
            url_tree=url_tree,
            attached_documents=attached_documents
        )

        # Verify metadata
        assert metadata.url == url
        assert "지하철 2호선" in metadata.title
        assert metadata.depth == 2
        assert metadata.word_count > 0
        assert len(metadata.entities_preview) > 0
        assert not metadata.has_attachments
        assert metadata.attachment_count == 0

    @pytest.mark.asyncio
    async def test_build_metadata_with_attachments(self, extractor, tmp_path):
        """Test metadata building with attachments"""
        # Create sample attachment
        attachment = AttachedDocument(
            attachment_id="test-001",
            attachment_type=AttachmentType.PDF,
            original_filename="report.pdf",
            file_size=1024,
            source_type=AttachmentSource.POPUP_WINDOW,
            source_url="http://example.com/report.pdf",
            popup_window_title="운행 보고서",
            popup_window_url="http://example.com/popup/report",
            local_path=str(tmp_path / "report.pdf"),
            extraction_method="pypdf",
            extraction_success=True,
            extracted_text="상세 운행 보고서",
            extracted_tables=[],
            extracted_images=[],
            word_count=50,
            entities_preview=["보고서", "운행"],
            named_entities={"TRANSPORT": ["2호선"]}
        )

        # Build metadata with attachment
        metadata = await extractor._build_metadata(
            url="https://news.seoul.go.kr/traffic/article/123",
            html="<html><body>Test</body></html>",
            markdown="Test content",
            links=[],
            url_tree={},
            attached_documents=[attachment]
        )

        # Verify attachment handling
        assert metadata.has_attachments is True
        assert metadata.attachment_count == 1
        assert AttachmentType.PDF in metadata.attachment_types
        assert metadata.total_attachment_size == 1024
        assert len(metadata.attached_documents) == 1

    def test_generate_extraction_report(self, extractor):
        """Test extraction report generation"""
        # Create sample metadata
        metadata_list = [
            PageMetadata(
                url="https://example.com/1",
                title="Page 1",
                page_type="ARTICLE",
                crawled_at=datetime.utcnow(),
                parent_url=None,
                depth=0,
                breadcrumb=[],
                siblings=[],
                outgoing_links=[],
                incoming_links=[],
                link_contexts=[],
                word_count=100,
                entities_preview=["test"],
                named_entities={},
                attached_documents=[],
                has_attachments=False,
                attachment_types=[],
                total_attachment_size=0,
                attachment_count=0
            ),
            PageMetadata(
                url="https://example.com/2",
                title="Page 2 with Attachment",
                page_type="ARTICLE",
                crawled_at=datetime.utcnow(),
                parent_url=None,
                depth=0,
                breadcrumb=[],
                siblings=[],
                outgoing_links=[],
                incoming_links=[],
                link_contexts=[],
                word_count=150,
                entities_preview=["test"],
                named_entities={},
                attached_documents=[
                    AttachedDocument(
                        attachment_id="a1",
                        attachment_type=AttachmentType.PDF,
                        original_filename="doc.pdf",
                        file_size=2048,
                        source_type=AttachmentSource.DIRECT_LINK,
                        source_url="http://example.com/doc.pdf",
                        local_path="/tmp/doc.pdf",
                        extraction_method="pypdf",
                        extraction_success=True,
                        extracted_text="test",
                        extracted_tables=[],
                        extracted_images=[],
                        word_count=10,
                        entities_preview=[],
                        named_entities={}
                    )
                ],
                has_attachments=True,
                attachment_types=[AttachmentType.PDF],
                total_attachment_size=2048,
                attachment_count=1
            )
        ]

        # Generate report
        report = extractor.generate_extraction_report(metadata_list)

        # Verify report content
        assert "Total Pages Extracted: 2" in report
        assert "Pages with Attachments: 1" in report
        assert "Total Attachments: 1" in report
        assert "ARTICLE:" in report
        assert "pdf:" in report


# Integration test (requires actual browser)
@pytest.mark.integration
@pytest.mark.asyncio
async def test_full_extraction_workflow():
    """
    Full integration test with actual crawling.

    Note: This test requires internet connection and may take time.
    Run with: pytest -m integration
    """
    import tempfile

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)

        extractor = SeoulTrafficContentExtractor(
            output_dir=tmp_path / "output",
            download_dir=tmp_path / "downloads",
            headless=True,
            verbose=False
        )

        # Test with a single known URL (adjust to actual test URL)
        test_urls = [
            {
                'url': 'https://news.seoul.go.kr/traffic',
                'priority': 1.0,
                'depth': 0
            }
        ]

        url_tree = {
            'https://news.seoul.go.kr/traffic': {
                'parent_url': None,
                'depth': 0,
                'children': []
            }
        }

        # Extract
        metadata_list = await extractor.extract_with_metadata(
            urls=test_urls,
            url_tree=url_tree,
            batch_size=1,
            delay_between_batches=0
        )

        # Verify results
        assert len(metadata_list) > 0
        assert metadata_list[0].url == test_urls[0]['url']
        assert (tmp_path / "output" / "raw").exists()
        assert (tmp_path / "output" / "metadata").exists()


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
